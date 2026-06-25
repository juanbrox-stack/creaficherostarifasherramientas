import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl.styles import Alignment, Font

# Configuración de la página de Streamlit
st.set_page_config(page_title="Actualizador de Tarifas e Ingestas", layout="wide", page_icon="📊")

# =========================================================================
# FUNCIONES AUXILIARES DE EXTRACCIÓN Y LIMPIEZA
# =========================================================================
def formatear_sku(val):
    if pd.isna(val):
        return ""
    val_str = str(val).strip()
    if val_str.endswith('.0'):
        val_str = val_str[:-2]
    
    if any(c.isalpha() for c in val_str):
        return val_str
        
    num_str = "".join(c for c in val_str if c.isdigit())
    if num_str:
        return num_str.zfill(5)
    return val_str

def limpiar_y_convertir_precio(val):
    if pd.isna(val):
        return None
    val_str = str(val).replace('€', '').replace('$', '').replace(' ', '').replace('zł', '').strip()
    if not val_str:
        return None
    
    if ',' in val_str and '.' in val_str:
        val_str = val_str.replace(',', '')
    elif ',' in val_str:
        val_str = val_str.replace(',', '.')
        
    try:
        return round(float(val_str), 2)
    except ValueError:
        return None

def extraer_precios_por_posicion(df, col_ref_idx, col_precio_idx):
    mapping = {}
    if df is None or df.empty:
        return mapping
        
    for index, row in df.iterrows():
        if col_ref_idx >= len(row) or pd.isna(row.iloc[col_ref_idx]):
            continue
            
        ref_raw = str(row.iloc[col_ref_idx]).strip()
        if ref_raw.upper() in ["REFERENCIA", "REFERENC", "SKU", "NOMBRE COMPLETO", "REFERENCE", "REF", ""]:
            continue
            
        sku = formatear_sku(ref_raw)
        if not sku:
            continue
            
        precio_val = None
        if col_precio_idx is not None and col_precio_idx < len(row):
            precio_val = limpiar_y_convertir_precio(row.iloc[col_precio_idx])
            
        mapping[sku] = precio_val
    return mapping

def exportar_caracteristica_excel(df_datos, nombre_caracteristica):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_datos.to_excel(writer, sheet_name='Sheet1', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        for row in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            if cell.value is not None:
                cell.number_format = '@'
                cell.value = str(cell.value)
                
    return output.getvalue()

def exportar_excel_con_cabecera_herramientas(df_datos, columnas_plantilla):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_datos.to_excel(writer, sheet_name='Sheet1', index=False, startrow=1)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        worksheet['A1'] = "Herramientas"
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas_plantilla))
        
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet['A1'].font = Font(bold=True, size=11)
        
        for row in range(3, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            if cell.value is not None:
                cell.number_format = '@'
                cell.value = str(cell.value)
                
    return output.getvalue()

def letra_columna(n):
    string = ""
    while n >= 0:
        n, remainder = divmod(n, 26)
        string = chr(65 + remainder) + string
        n -= 1
    return string

# =========================================================================
# BARRA LATERAL - CARGA DE ARCHIVOS MAESTROS Y VALIDACIÓN SINCRO
# =========================================================================
st.sidebar.header("📂 Archivos Maestros de Origen")
archivo_nac = st.sidebar.file_uploader("1. Tarifa Nacional (Excel)", type=["xlsx"])
archivo_int = st.sidebar.file_uploader("2. Tarifa Internacional (Excel)", type=["xlsx"])
archivo_vent = st.sidebar.file_uploader("3. Tarifa Internacional VENT. (Excel)", type=["xlsx"])

pestañas_nac = {}
pestañas_int = {}
pestañas_vent = {}

pestañas_nac_esperadas = ["AMAZON", "MIRAVIA", "MEDIAMARKT", "CARREFOUR", "PRIVALIA"]
pestañas_int_esperadas = [
    "FRANCIA (ES-FR)", "FRANCIA (FR-FR)", "ITALIA (ES-IT)", "ITALIA (IT-IT)", 
    "ALEMANAI (ES-DE)", "ALEMANIA (DE-DE)", "PORTUGAL", "HOLANDA", "BELGICA", "POLONIA", "SUECIA"
]
pestañas_vent_esperadas = [
    "VENT. ES-FR", "VENT. FR-FR", "VENT. ES-IT", "VENT. IT-IT",
    "VENT. ES-DE", "VENT. DE-DE"
]

if archivo_nac:
    try:
        xl_nac = pd.ExcelFile(archivo_nac)
        pestañas_detectadas = xl_nac.sheet_names
        for sheet in pestañas_detectadas:
            pestañas_nac[sheet.strip()] = xl_nac.parse(sheet, header=None)
        
        pestañas_faltantes = [p for p in pestañas_nac_esperadas if p not in pestañas_detectadas]
        pestañas_nuevas = [p for p in pestañas_detectadas if p not in pestañas_nac_esperadas]
        
        if pestañas_faltantes:
            st.sidebar.error(f"⚠️ Faltan pestañas críticas Nacionales: {pestañas_faltantes}")
        if pestañas_nuevas:
            st.sidebar.warning(f"💡 Pestañas adicionales detectadas: {pestañas_nuevas}")
        if not pestañas_faltantes:
            st.sidebar.success("✅ Estructura Nacional validada correctamente.")
    except Exception as e:
        st.sidebar.error(f"Error al leer Tarifa Nacional: {e}")

if archivo_int:
    try:
        xl_int = pd.ExcelFile(archivo_int)
        pestañas_detectadas_int = xl_int.sheet_names
        for sheet in pestañas_detectadas_int:
            pestañas_int[sheet.strip()] = xl_int.parse(sheet, header=None)
            
        pestañas_faltantes_int = [p for p in pestañas_int_esperadas if p not in pestañas_detectadas_int]
        pestañas_nuevas_int = [p for p in pestañas_detectadas_int if p not in pestañas_int_esperadas]
        
        if pestañas_faltantes_int:
            st.sidebar.error(f"⚠️ Faltan pestañas críticas Internacionales: {pestañas_faltantes_int}")
        if pestañas_nuevas_int:
            st.sidebar.warning(f"💡 Pestañas adicionales detectadas: {pestañas_nuevas_int}")
        if not pestañas_faltantes_int:
            st.sidebar.success("✅ Estructura Internacional validada correctamente.")
    except Exception as e:
        st.sidebar.error(f"Error al leer Tarifa Internacional: {e}")

if archivo_vent:
    try:
        xl_vent = pd.ExcelFile(archivo_vent)
        pestañas_detectadas_vent = xl_vent.sheet_names
        for sheet in pestañas_detectadas_vent:
            nombre_limpio = sheet.strip()
            pestañas_vent[nombre_limpio] = xl_vent.parse(sheet, header=None)

        # Solo mostramos las pestanas encontradas; no hay obligatorias
        nombres_limpios_vent = list(pestañas_vent.keys())
        encontradas_vent = [p for p in nombres_limpios_vent if p in pestañas_vent_esperadas]
        extras_vent = [p for p in nombres_limpios_vent if p not in pestañas_vent_esperadas]

        if encontradas_vent:
            st.sidebar.success(f"✅ VENT. cargado: {encontradas_vent}")
        if extras_vent:
            st.sidebar.info(f"💡 Otras pestañas detectadas: {extras_vent}")
        if not encontradas_vent and not extras_vent:
            st.sidebar.warning("⚠️ No se encontraron pestañas reconocidas en el archivo VENT.")
    except Exception as e:
        st.sidebar.error(f"Error al leer Tarifa VENT.: {e}")

# =========================================================================
# PANEL DE CONTROL DINÁMICO DE MAPEO (VALORES POR DEFECTO ACTUALIZADOS)
# =========================================================================
st.write("---")
with st.expander("🛠️ Panel Avanzado de Mapeo y Remapeo de Columnas"):
    st.write("Configuración de columnas por defecto actualizada según la nueva estructura del Excel:")
    c_map1, c_map2, c_map3 = st.columns(3)
    
    with c_map1:
        st.markdown("**📌 Canales Nacionales**")
        sample_nac = pestañas_nac.get("AMAZON") if "AMAZON" in pestañas_nac else (next(iter(pestañas_nac.values())) if pestañas_nac else None)
        max_cols_nac = len(sample_nac.columns) if sample_nac is not None else 24
        opciones_cols_nac = [f"Columna {letra_columna(i)} (Índice {i})" for i in range(max_cols_nac)]
        
        # Puesta por defecto la Columna A (Índice 0) para Referencia y Columna R (Índice 17) para PVP PUBLICADO
        re_ref_nac = st.selectbox("Columna de Referencia / SKU (Nacional):", opciones_cols_nac, index=0)
        re_pvp_nac = st.selectbox("Columna PVP PUBLICADO unificado (Nacional):", opciones_cols_nac, index=17)
        
        idx_ref_nac = opciones_cols_nac.index(re_ref_nac)
        idx_pvp_nac_def = opciones_cols_nac.index(re_pvp_nac)

    with c_map2:
        st.markdown("**📌 Canales Internacionales**")
        sample_int = pestañas_int.get("FRANCIA (ES-FR)") if "FRANCIA (ES-FR)" in pestañas_int else (next(iter(pestañas_int.values())) if pestañas_int else None)
        max_cols_int = len(sample_int.columns) if sample_int is not None else 24
        opciones_cols_int = [f"Columna {letra_columna(i)} (Índice {i})" for i in range(max_cols_int)]
        
        # Valores de referencia internacionales configurables dinámicamente si varía el Excel
        re_ref_int = st.selectbox("Columna de Referencia / SKU (Internacional):", opciones_cols_int, index=0)
        re_pvp_es_xxx = st.selectbox("Columna PVP para Cruces ES-XX (Columna S):", opciones_cols_int, index=18)
        re_pvp_xx_xx  = st.selectbox("Columna PVP para Cruces XX-XX (Columna T):", opciones_cols_int, index=19)
        re_pvp_std_int = st.selectbox("Columna PVP Estándar PT, NL, BE (Columna R):", opciones_cols_int, index=17)
        re_pvp_div_int = st.selectbox("Columna PVPR Divisas Especiales PL, SE (Columna I):", opciones_cols_int, index=8)
        
        idx_ref_int = opciones_cols_int.index(re_ref_int)
        idx_s_precio = opciones_cols_int.index(re_pvp_es_xxx)
        idx_t_precio = opciones_cols_int.index(re_pvp_xx_xx)
        idx_r_precio = opciones_cols_int.index(re_pvp_std_int)
        idx_i_precio = opciones_cols_int.index(re_pvp_div_int)

    with c_map3:
        st.markdown("**📌 Canales VENT. (Nueva Tarifa Internacional)**")
        sample_vent = next(iter(pestañas_vent.values())) if pestañas_vent else None
        max_cols_vent = len(sample_vent.columns) if sample_vent is not None else 24
        opciones_cols_vent = [f"Columna {letra_columna(i)} (Índice {i})" for i in range(max_cols_vent)]

        re_ref_vent = st.selectbox("Columna de Referencia / SKU (VENT.):", opciones_cols_vent, index=0)
        re_pvp_vent = st.selectbox("Columna PVP PUB (VENT.) — por defecto Col. L:", opciones_cols_vent, index=min(11, max_cols_vent - 1))

        idx_ref_vent = opciones_cols_vent.index(re_ref_vent)
        idx_pvp_vent = opciones_cols_vent.index(re_pvp_vent)

        if sample_vent is not None:
            st.markdown("**Preview primeras 5 filas del archivo VENT.:**")
            preview = sample_vent.head(5).copy()
            preview.columns = [f"{letra_columna(i)} ({i})" for i in range(len(preview.columns))]
            st.dataframe(preview, use_container_width=True)

columnas_plantilla = [
    'reference', 'price_france', 'price_italy', 'price_germany', 'price_portugal',
    'price_spain', 'price_poland', 'price_holand', 'price_tradeinn_es', 'price_aliexpress_es',
    'price_makro_es', 'price_mediamarkt_es', 'price_aurgi_es', 'price_elcorteingles_es',
    'price_makro_de', 'price_makro_it', 'price_carrefour', 'price_pccomponentes'
]
columnas_de_precio_totales = [c for c in columnas_plantilla if c != 'reference']

# =========================================================================
# VISTA DE PESTAÑAS DE TRABAJO
# =========================================================================
tab1, tab2 = st.tabs(["📦 Bloque 1: Características", "🚀 Bloque 2: Cargador de Precios (3 Ficheros)"])

# -------------------------------------------------------------------------
# BLOQUE 1: PROCESAMIENTO MASIVO POR CARACTERÍSTICAS
# -------------------------------------------------------------------------
with tab1:
    st.header("Generación Masiva por Características")
    st.write("Genera los archivos individuales organizados y limpios en formato **Excel (.xlsx)**.")

    reglas_caracteristicas = {
        # España (Tarifa Nacional unificada a Columna R)
        "PVP_LEROYES": ("Nacional", ["AMAZON"], idx_ref_nac, idx_pvp_nac_def), 
        "PVP_PcComponentes": ("Nacional", ["MEDIAMARKT"], idx_ref_nac, idx_pvp_nac_def),
        "PVPR": ("Nacional", ["AMAZON"], idx_ref_nac, idx_pvp_nac_def),
        "PVP ESPANA": ("Nacional", ["AMAZON"], idx_ref_nac, idx_pvp_nac_def),
        "Pvp mediamarkt es": ("Nacional", ["MEDIAMARKT"], idx_ref_nac, idx_pvp_nac_def),
        "PVP_SHEIN_ES": ("Nacional", ["AMAZON"], idx_ref_nac, idx_pvp_nac_def),
        # Internacional estructurado por columnas explícitas
        "Prix_France": ("Internacional", ["FRANCIA (ES-FR)"], idx_ref_int, idx_s_precio),
        "PVP_SHEIN_FR": ("Internacional", ["FRANCIA (ES-FR)"], idx_ref_int, idx_s_precio),
        "Prix_Italia": ("Internacional", ["ITALIA (ES-IT)"], idx_ref_int, idx_s_precio),
        "PVP_SHEIN_IT": ("Internacional", ["ITALIA (ES-IT)"], idx_ref_int, idx_s_precio),
        "prix_Alemania": ("Internacional", ["ALEMANAI (ES-DE)"], idx_ref_int, idx_s_precio),
        "PVP_SHEIN_DE": ("Internacional", ["ALEMANAI (ES-DE)"], idx_ref_int, idx_s_precio),
        "PVP_SHEIN_PT": ("Internacional", ["PORTUGAL"], idx_ref_int, idx_r_precio),
        "PVP PORTUGAL": ("Internacional", ["PORTUGAL"], idx_ref_int, idx_r_precio),
        "PVP_BE": ("Internacional", ["BELGICA"], idx_ref_int, idx_r_precio),
        "PRIXHOLANDA": ("Internacional", ["HOLANDA"], idx_ref_int, idx_r_precio),
        "PVP_SHEIN_NL": ("Internacional", ["HOLANDA"], idx_ref_int, idx_r_precio),
        "PVP_SHEIN_PL": ("Internacional", ["POLONIA"], idx_ref_int, idx_i_precio),
        "PRIX_POLONIA": ("Internacional", ["POLONIA"], idx_ref_int, idx_i_precio),
        "PVP Suecia": ("Internacional", ["SUECIA"], idx_ref_int, idx_i_precio),
        # VENT. — Nueva Tarifa Internacional (Col. L = índice 11 por defecto)
        "PVP_VENT_ES-FR": ("VENT", ["VENT. ES-FR"], idx_ref_vent, idx_pvp_vent),
        "PVP_VENT_FR-FR": ("VENT", ["VENT. FR-FR"], idx_ref_vent, idx_pvp_vent),
        "PVP_VENT_ES-IT": ("VENT", ["VENT. ES-IT"], idx_ref_vent, idx_pvp_vent),
        "PVP_VENT_IT-IT": ("VENT", ["VENT. IT-IT"], idx_ref_vent, idx_pvp_vent),
        "PVP_VENT_ES-DE": ("VENT", ["VENT. ES-DE"], idx_ref_vent, idx_pvp_vent),
        "PVP_VENT_DE-DE": ("VENT", ["VENT. DE-DE"], idx_ref_vent, idx_pvp_vent),
    }

    if st.button("📦 Generar todas las Características"):
        if not archivo_nac and not archivo_int and not archivo_vent:
            st.error("Por favor, sube al menos un archivo maestro en la barra lateral.")
        else:
            diccionario_archivos = {}
            for nombre_carac, (tipo, posibles_pestañas, col_ref, col_precio) in reglas_caracteristicas.items():
                df_trabajo = None
                mapping_precios = {}
                
                if tipo == "Nacional":
                    df_trabajo = pestañas_nac.get(posibles_pestañas[0])
                    mapping_precios = extraer_precios_por_posicion(df_trabajo, col_ref, col_precio)
                elif tipo == "Internacional":
                    for p in posibles_pestañas:
                        if p in pestañas_int:
                            df_trabajo = pestañas_int[p]
                            break
                    mapping_precios = extraer_precios_por_posicion(df_trabajo, col_ref, col_precio)
                elif tipo == "VENT":
                    if not archivo_vent:
                        continue  # saltar si no se ha subido la tarifa VENT
                    for p in posibles_pestañas:
                        if p in pestañas_vent:
                            df_trabajo = pestañas_vent[p]
                            break
                    mapping_precios = extraer_precios_por_posicion(df_trabajo, col_ref, col_precio)
                
                if mapping_precios:
                    df_out = pd.DataFrame(list(mapping_precios.items()), columns=['sku', 'valor'])
                    df_out = df_out[df_out['valor'].notna()]
                    df_out['valor'] = df_out['valor'].apply(lambda x: "{:.2f}".format(x) if isinstance(x, (int, float)) else str(x))
                    df_out = df_out[df_out['valor'] != ""]
                    
                    if not df_out.empty:
                        diccionario_archivos[nombre_carac] = exportar_caracteristica_excel(df_out, nombre_carac)

            if diccionario_archivos:
                st.session_state["archivos_caracteristicas"] = diccionario_archivos
                st.success(f"¡Procesamiento completo! {len(diccionario_archivos)} archivos preparados.")
            else:
                st.error("No se pudo extraer información válida. Verifica las columnas seleccionadas.")

    if "archivos_caracteristicas" in st.session_state:
        archivos = st.session_state["archivos_caracteristicas"]
        st.write("### ⬇️ Descarga de Ficheros Individuales (.xlsx)")
        cols = st.columns(3)
        for idx, (nombre_fichero, datos_excel) in enumerate(archivos.items()):
            col_actual = cols[idx % 3]
            with col_actual:
                st.download_button(
                    label=f"📊 {nombre_fichero}.xlsx", data=datos_excel,
                    file_name=f"{nombre_fichero}.xlsx", key=f"btn_{nombre_fichero}",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        st.write("---")
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for nombre_fichero, datos_excel in archivos.items():
                zip_file.writestr(f"{nombre_fichero}.xlsx", datos_excel)
        st.download_button(
            label="📥 Descargar TODOS los archivos de Características (.ZIP)",
            data=zip_buffer.getvalue(), file_name="caracteristicas_actualizadas_excel.zip",
            mime="application/zip", type="primary"
        )

# -------------------------------------------------------------------------
# BLOQUE 2: CARGADOR DE PRECIOS
# -------------------------------------------------------------------------
with tab2:
    st.header("Generación del Cargador de Precios")
    st.write("Presiona el botón para procesar las tarifas horizontales.")

    if st.button("🚀 Procesar y Preparar Ficheros del Cargador"):
        if not archivo_nac and not archivo_int and not archivo_vent:
            st.error("Sube al menos un archivo maestro en la barra lateral.")
        else:
            cargador_generado = {}

            def fmt_precio(x):
                return f"{x:.2f}" if isinstance(x, (int, float)) else x

            def df_limpio(df_in):
                df_in = df_in.dropna(subset=columnas_de_precio_totales, how='all').fillna("")
                for c in columnas_plantilla:
                    if c != 'reference':
                        df_in[c] = df_in[c].apply(fmt_precio)
                return df_in

            # --- FICHERO 1: ESPAÑA ---
            df_amz = pestañas_nac.get("AMAZON")
            if df_amz is not None:
                map_amz = extraer_precios_por_posicion(df_amz, idx_ref_nac, idx_pvp_nac_def)
                map_mir = extraer_precios_por_posicion(pestañas_nac.get("MIRAVIA"), idx_ref_nac, idx_pvp_nac_def)
                map_mm  = extraer_precios_por_posicion(pestañas_nac.get("MEDIAMARKT"), idx_ref_nac, idx_pvp_nac_def)
                map_c4  = extraer_precios_por_posicion(pestañas_nac.get("CARREFOUR"), idx_ref_nac, idx_pvp_nac_def)
                df_f1 = pd.DataFrame(columns=columnas_plantilla)
                df_f1['reference'] = list(map_amz.keys())
                df_f1['price_spain']        = df_f1['reference'].map(map_amz)
                df_f1['price_tradeinn_es']  = df_f1['reference'].map(map_amz)
                df_f1['price_aliexpress_es']= df_f1['reference'].map(map_mir)
                df_f1['price_mediamarkt_es']= df_f1['reference'].map(map_mm)
                df_f1['price_aurgi_es']     = df_f1['reference'].map(map_amz)
                df_f1['price_carrefour']    = df_f1['reference'].map(map_c4)
                df_f1['price_pccomponentes']= df_f1['reference'].map(map_mm)
                cargador_generado["f1"] = exportar_excel_con_cabecera_herramientas(df_limpio(df_f1), columnas_plantilla)
            else:
                cargador_generado["f1"] = None

            # --- FICHERO 2: PORTUGAL ---
            df_pt = pestañas_int.get("PORTUGAL")
            if df_pt is not None:
                map_pt = extraer_precios_por_posicion(df_pt, idx_ref_int, idx_r_precio)
                df_f2 = pd.DataFrame(columns=columnas_plantilla)
                df_f2['reference'] = list(map_pt.keys())
                df_f2['price_portugal'] = list(map_pt.values())
                cargador_generado["f2"] = exportar_excel_con_cabecera_herramientas(df_limpio(df_f2), columnas_plantilla)
            else:
                cargador_generado["f2"] = None

            # --- FICHERO 3: RESTO EUROPA (Internacional + VENT. ES-XX) ---
            def buscar_y_extraer_int(lista_pestañas, col_ref, col_precio):
                for p in lista_pestañas:
                    if p in pestañas_int:
                        return extraer_precios_por_posicion(pestañas_int[p], col_ref, col_precio)
                return {}

            def buscar_y_extraer_vent(pestaña):
                df_v = pestañas_vent.get(pestaña)
                return extraer_precios_por_posicion(df_v, idx_ref_vent, idx_pvp_vent) if df_v is not None else {}

            # Internacional
            map_fr = buscar_y_extraer_int(["FRANCIA (ES-FR)"], idx_ref_int, idx_s_precio)
            map_it = buscar_y_extraer_int(["ITALIA (ES-IT)"], idx_ref_int, idx_s_precio)
            map_de = buscar_y_extraer_int(["ALEMANAI (ES-DE)"], idx_ref_int, idx_s_precio)
            map_nl = buscar_y_extraer_int(["HOLANDA"], idx_ref_int, idx_r_precio)
            map_be = buscar_y_extraer_int(["BELGICA"], idx_ref_int, idx_r_precio)
            map_pl = buscar_y_extraer_int(["POLONIA"], idx_ref_int, idx_i_precio)

            # VENT. ES-XX: si hay datos VENT, se usan para price_france/italy/germany
            # complementando (combine_first) los datos del Internacional
            if archivo_vent and pestañas_vent:
                map_vent_fr = buscar_y_extraer_vent("VENT. ES-FR")
                map_vent_it = buscar_y_extraer_vent("VENT. ES-IT")
                map_vent_de = buscar_y_extraer_vent("VENT. ES-DE")
            else:
                map_vent_fr = {}
                map_vent_it = {}
                map_vent_de = {}

            all_refs_f3 = sorted(set(
                list(map_fr) + list(map_it) + list(map_de) +
                list(map_nl) + list(map_be) + list(map_pl) +
                list(map_vent_fr) + list(map_vent_it) + list(map_vent_de)
            ))

            if all_refs_f3:
                df_f3 = pd.DataFrame(columns=columnas_plantilla)
                df_f3['reference'] = all_refs_f3

                # price_france: Internacional primero, VENT. ES-FR como fallback (o al revés si no hay int)
                serie_fr_int  = df_f3['reference'].map(map_fr)
                serie_fr_vent = df_f3['reference'].map(map_vent_fr)
                df_f3['price_france']  = serie_fr_int.combine_first(serie_fr_vent)

                serie_it_int  = df_f3['reference'].map(map_it)
                serie_it_vent = df_f3['reference'].map(map_vent_it)
                df_f3['price_italy']   = serie_it_int.combine_first(serie_it_vent)

                serie_de_int  = df_f3['reference'].map(map_de)
                serie_de_vent = df_f3['reference'].map(map_vent_de)
                df_f3['price_germany'] = serie_de_int.combine_first(serie_de_vent)

                df_f3['price_holand'] = df_f3['reference'].map(map_nl).combine_first(df_f3['reference'].map(map_be))
                df_f3['price_poland'] = df_f3['reference'].map(map_pl)

                cargador_generado["f3"] = exportar_excel_con_cabecera_herramientas(df_limpio(df_f3), columnas_plantilla)
            else:
                cargador_generado["f3"] = None

            cargador_generado["f4"] = None  # ya no existe fichero 4 separado

            st.session_state["cargador_generado"] = cargador_generado
            n_ok = sum(1 for v in cargador_generado.values() if v is not None)
            st.success(f"¡Procesamiento completo! {n_ok} fichero(s) generado(s).")

if "cargador_generado" in st.session_state:
    cg = st.session_state["cargador_generado"]
    st.write("### ⬇️ Descarga de Plantillas Horizontales (.xlsx)")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.subheader("1. España")
        if cg.get("f1"):
            st.download_button("📥 Tarifa Nacional", data=cg["f1"], file_name="1_Tarifa_Nacional_Espana.xlsx")
        else:
            st.caption("Sin datos (sube Tarifa Nacional).")
    with col2:
        st.subheader("2. Portugal")
        if cg.get("f2"):
            st.download_button("📥 Tarifa Portugal", data=cg["f2"], file_name="2_Tarifa_Internacional_Portugal.xlsx")
        else:
            st.caption("Sin datos (sube Tarifa Internacional con pestaña PORTUGAL).")
    with col3:
        st.subheader("3. Resto Europa (+ VENT. ES-XX)")
        if cg.get("f3"):
            st.download_button("📥 Resto Europa", data=cg["f3"], file_name="3_Resto_de_Internacional.xlsx")
        else:
            st.caption("Sin datos (sube Tarifa Internacional y/o VENT.).")

    st.write("---")
    st.write("### 🗜️ Descarga Conjunta del Cargador")
    zip_cargador_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_cargador_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        nombres = {
            "f1": "1_Tarifa_Nacional_Espana.xlsx",
            "f2": "2_Tarifa_Internacional_Portugal.xlsx",
            "f3": "3_Resto_de_Internacional.xlsx",
        }
        for key, nombre in nombres.items():
            if cg.get(key):
                zf.writestr(nombre, cg[key])
    st.download_button(
        label="📦 Descargar todos los ficheros del Cargador (.ZIP)",
        data=zip_cargador_buffer.getvalue(), file_name="cargador_tarifas_completo.zip",
        mime="application/zip", type="primary"
    )
